from openpyxl import load_workbook
import os

def analyze_original_file():
    """
    Analisa o arquivo original para entender sua estrutura e formatação
    """
    try:
        # Caminho para o arquivo original
        base_dir = r"c:\Users\palad\OneDrive\Documentos\Operacional\PROGRAMA DE CÁLCULOS\Django\HTcalculus"
        excel_path = os.path.join(base_dir, "Descompressao", "Cálculos HT 2.0 - Descompressão.xlsm")
        
        if not os.path.exists(excel_path):
            print(f"Arquivo não encontrado: {excel_path}")
            return
            
        # Carregar o arquivo original
        wb = load_workbook(filename=excel_path, read_only=False, keep_vba=False)  # Changed to read_only=False
        ws = wb.active
        
        print(f"Nome da planilha: {ws.title}")
        print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
        print("\n" + "="*50)
        
        # Analisar as primeiras 20 linhas
        print("ESTRUTURA DAS PRIMEIRAS 20 LINHAS:")
        print("="*50)
        
        for row in range(1, 21):
            line_content = []
            for col in range(1, 25):  # Todas as 24 colunas (A até X)
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                    line_content.append(f"{col_letter}({col}): {cell.value}")
            
            if line_content:
                print(f"Linha {row}: {' | '.join(line_content)}")
        
        print("\n" + "="*50)
        print("ANÁLISE DETALHADA DE CÉLULAS ESPECÍFICAS:")
        print("="*50)
        
        # Células importantes baseadas no seu código
        important_cells = ['A1', 'D1', 'A2', 'D2', 'A6', 'A7', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'P7', 'Q7', 'R7', 'S7', 'T7', 'U7', 'V7', 'W7', 'X7',
                           'A8', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'P8', 'Q8', 'R8', 'S8', 'T8', 'U8', 'V8', 'W8', 'X8']
        
        for cell_ref in important_cells:
            cell = ws[cell_ref]
            if cell.value is not None:
                print(f"{cell_ref}: '{cell.value}' | Font: {cell.font.name if cell.font else 'N/A'} | "
                      f"Size: {cell.font.size if cell.font else 'N/A'} | "
                      f"Bold: {cell.font.bold if cell.font else 'N/A'}")
        
        print("\n" + "="*50)
        print("DATAS NAS LINHAS 10-20:")
        print("="*50)
        
        for row in range(10, 21):
            cell_a = ws.cell(row=row, column=1)
            if cell_a.value:
                print(f"Linha {row}, Col A: {cell_a.value} | Formato: {cell_a.number_format}")
        
        print("\n" + "="*50)
        print("ANÁLISE DE CORES E BORDAS (primeiras células):")
        print("="*50)
        
        for row in range(1, 15):
            for col in range(1, 25):
                cell = ws.cell(row=row, column=col)
                col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                
                # Verificar cores de fundo
                if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000' and cell.fill.start_color.index != 'FFFFFFFF':
                    print(f"Célula {col_letter}{row}: Cor de fundo: {cell.fill.start_color.index} | Fill type: {cell.fill.fill_type}")
                
                # Verificar cores de fonte
                if cell.font and cell.font.color and cell.font.color.index != '00000000':
                    print(f"Célula {col_letter}{row}: Cor da fonte: {cell.font.color.index}")
                
                # Verificar bordas
                if cell.border and any([cell.border.left.style, cell.border.right.style, 
                                      cell.border.top.style, cell.border.bottom.style]):
                    border_info = []
                    if cell.border.left.style: border_info.append(f"left:{cell.border.left.style}")
                    if cell.border.right.style: border_info.append(f"right:{cell.border.right.style}")
                    if cell.border.top.style: border_info.append(f"top:{cell.border.top.style}")
                    if cell.border.bottom.style: border_info.append(f"bottom:{cell.border.bottom.style}")
                    if border_info and row <= 10:  # Só mostrar bordas das primeiras linhas
                        print(f"Célula {col_letter}{row}: Bordas: {', '.join(border_info)}")
        
        print("\n" + "="*50)
        print("ANÁLISE ESPECÍFICA DE CORES NAS LINHAS DE DADOS:")
        print("="*50)
        
        # Analisar algumas linhas de dados para ver padrões de cores
        for row in range(10, 25):
            for col in range(1, 25):
                cell = ws.cell(row=row, column=col)
                col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                
                if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000' and cell.fill.start_color.index != 'FFFFFFFF':
                    print(f"Célula {col_letter}{row}: Cor: {cell.fill.start_color.index} | Valor: {cell.value}")
        
        print("\n" + "="*50)
        print("ANÁLISE DE CÉLULAS MESCLADAS:")
        print("="*50)
        
        # Verificar células mescladas
        for merged_range in ws.merged_cells.ranges:
            print(f"Células mescladas: {merged_range}")
            # Pegar o valor da célula superior esquerda
            top_left = ws[str(merged_range).split(':')[0]]
            print(f"  Valor: '{top_left.value}' | Font: {top_left.font.name if top_left.font else 'N/A'} | "
                  f"Size: {top_left.font.size if top_left.font else 'N/A'} | Bold: {top_left.font.bold if top_left.font else 'N/A'}")
        
        wb.close()
        
    except Exception as e:
        print(f"Erro ao analisar arquivo: {e}")

if __name__ == "__main__":
    analyze_original_file()
