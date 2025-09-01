from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

def criar_planilha_descompressao_original():
    """
    Cria uma planilha Excel exatamente igual ao arquivo original 'Cálculos HT 2.0 - Descompressão.xlsm'
    baseado na análise do arquivo original
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ModeloDescNovo"  # Nome da planilha original
    
    # Definir estilos baseados no arquivo original
    header_font = Font(name='Arial', size=8, bold=True)
    normal_font = Font(name='Arial', size=8, bold=False)
    
    # Cores específicas identificadas no arquivo original
    blue_font = Font(name='Arial', size=8, bold=False, color='FF2A0BE7')  # Azul usado em várias células
    gray_background = PatternFill(start_color='FFD0D0D0', end_color='FFD0D0D0', fill_type='solid')  # Fundo cinza para headers R7-R9
    light_blue_background = PatternFill(start_color='FFCCFFFF', end_color='FFCCFFFF', fill_type='solid')  # Fundo azul claro para colunas R,T,V,W
    
    # Alinhamento para células mescladas
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Bordas baseadas no arquivo original
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # === CABEÇALHO PRINCIPAL (Linhas 1-4) ===
    # Linha 1
    ws['A1'] = 'NOME'
    ws['A1'].font = header_font
    ws['A1'].border = thin_border
    
    ws['K1'] = 'CITAÇÃO'
    ws['K1'].font = header_font
    ws['K1'].border = thin_border
    
    ws['L1'] = datetime(2000, 11, 20)  # Data padrão do original
    ws['L1'].font = normal_font
    ws['L1'].border = thin_border
    ws['L1'].number_format = 'dd/mm/yyyy'
    
    ws['N1'] = 'JUROS'
    ws['N1'].font = header_font
    ws['N1'].border = thin_border
    
    ws['O1'] = 'POUPANÇA + SELIC'
    ws['O1'].font = header_font
    ws['O1'].border = thin_border
    
    # Linha 2
    ws['A2'] = 'MATRÍCULA'
    ws['A2'].font = header_font
    ws['A2'].border = thin_border
    
    ws['I2'] = 'HON SUCUM'
    ws['I2'].font = header_font
    ws['I2'].border = thin_border
    
    ws['J2'] = 0.05
    ws['J2'].font = normal_font
    ws['J2'].border = thin_border
    
    ws['K2'] = 'ATUALIZ.'
    ws['K2'].font = header_font
    ws['K2'].border = thin_border
    
    ws['L2'] = 'IPCA'
    ws['L2'].font = header_font
    ws['L2'].border = thin_border
    
    ws['Q2'] = '=T1-L1'  # Fórmula do original
    ws['Q2'].font = normal_font
    ws['Q2'].border = thin_border
    
    # Linha 3
    ws['A3'] = 'MÊS DE INÍCIO'
    ws['A3'].font = header_font
    ws['A3'].border = thin_border
    
    ws['I3'] = 'NOMEAÇÃO'
    ws['I3'].font = header_font
    ws['I3'].border = thin_border
    
    # Linha 4
    ws['A4'] = 'CPF'
    ws['A4'].font = header_font
    ws['A4'].border = thin_border
    
    ws['E4'] = 'ENDEREÇO'
    ws['E4'].font = header_font
    ws['E4'].border = thin_border
    
    # === TÍTULO DO RELATÓRIO (Linha 6) ===
    ws['A6'] = 'Relatório de Atualização Mensal -'
    ws['A6'].font = header_font
    ws['A6'].border = medium_border  # Bordas médias como no original
    
    # === CABEÇALHOS DAS COLUNAS (Linhas 7-8) ===
    # Linha 7 - Cabeçalhos principais
    headers_row7 = {
        'A7': 'Mês',
        'B7': 'Nivel', 
        'C7': 'Vencimento',
        'F7': 'GAM',
        'J7': 'Titulação',
        'L7': 'GCET',
        'N7': 'Adic. Tempo Serv.',
        'P7': 'Férias e 13° Salário',
        'R7': 'Diferença Total',
        'S7': 'Índice de Atualização',
        'T7': 'Valor corrigido',
        'U7': 'Id.Juros (%)',
        'V7': 'Juros de Mora',
        'W7': 'Principal + juros'
    }
    
    for cell, value in headers_row7.items():
        ws[cell] = value
        ws[cell].font = header_font
        ws[cell].border = medium_border  # Bordas médias para headers principais
    
    # Linha 8 - Sub-cabeçalhos
    headers_row8 = {
        'C8': 'Pago',
        'D8': 'Devido', 
        'E8': 'Diferença',
        'F8': '%',
        'G8': 'Pago',
        'H8': 'Devido',
        'I8': 'Diferença',
        'J8': '%',
        'K8': 'Diferença',
        'L8': '%',
        'M8': 'Diferença',
        'N8': '%',
        'O8': 'Diferença',
        'P8': '%',
        'Q8': 'Diferença'
    }
    
    for cell, value in headers_row8.items():
        ws[cell] = value
        ws[cell].font = header_font
        ws[cell].border = medium_border  # Bordas médias para sub-headers
    
    # Linha 9 - Explicações das fórmulas
    headers_row9 = {
        'A9': 'Referência',
        'R9': '(a)',
        'S9': '(b)',
        'T9': '(c = a x b)',
        'U9': '(d)',
        'V9': '(e = c x d)',
        'W9': '(f = c + e)'
    }
    
    for cell, value in headers_row9.items():
        ws[cell] = value
        ws[cell].font = header_font
        ws[cell].border = thin_border
    
    # === DADOS (Linhas 10-91) ===
    # Gerar datas de fev/1998 a nov/2004 (82 linhas)
    start_date = datetime(1998, 2, 1)
    
    for row in range(10, 92):  # 82 linhas de dados (10 a 91)
        months_passed = row - 10
        
        # Calcular a data correta
        month = 2 + months_passed  # Começar em fevereiro
        year = 1998
        
        while month > 12:
            month -= 12
            year += 1
            
        date_value = datetime(year, month, 1)
        
        # Inserir data na coluna A
        ws.cell(row=row, column=1).value = date_value
        ws.cell(row=row, column=1).number_format = 'mmm-yy'  # Formato original
        ws.cell(row=row, column=1).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        
        # Adicionar bordas e valores padrão nas outras colunas (até coluna X = 24)
        for col in range(2, 25):  # 24 colunas no total
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            
            # Valores padrão para algumas colunas importantes
            if col in [3, 7, 10]:  # Colunas C, G, J (valores pagos)
                cell.value = 0
    
    # === CÉLULAS MESCLADAS ===
    # Baseado na análise do arquivo original
    merged_cells_ranges = [
        'A1:C1',    # NOME
        'A2:C2',    # MATRÍCULA  
        'A3:C3',    # MÊS DE INÍCIO
        'O1:Q1',    # POUPANÇA + SELIC
        'U1:V1',    # Célula vazia
        'U2:V2',    # Célula vazia
        'J6:S6',    # Célula vazia no título
        'A7:A8',    # Mês
        'B7:B9',    # Nivel
        'C7:E7',    # Vencimento
        'F7:I7',    # GAM
        'J7:K7',    # Titulação
        'L7:M7',    # GCET
        'N7:O7',    # Adic. Tempo Serv.
        'P7:Q7',    # Férias e 13° Salário
        'R7:R8',    # Diferença Total
        'S7:S8',    # Índice de Atualização
        'T7:T8',    # Valor corrigido
        'U7:U8',    # Id.Juros (%)
        'V7:V8',    # Juros de Mora
        'W7:W8',    # Principal + juros
        'C8:C9',    # Vencimento Pago
        'D8:D9',    # Vencimento Devido
        'E8:E9',    # Vencimento Diferença
        'F8:F9',    # GAM %
        'G8:G9',    # GAM Pago
        'H8:H9',    # GAM Devido
        'I8:I9',    # GAM Diferença
        'J8:J9',    # Titulação %
        'K8:K9',    # Titulação Diferença
        'L8:L9',    # GCET %
        'M8:M9',    # GCET Diferença
        'N8:N9',    # Adic. Tempo Serv. %
        'O8:O9',    # Adic. Tempo Serv. Diferença
        'P8:P9',    # Férias %
        'Q8:Q9',    # Férias Diferença
        'A92:S92',  # VALOR DA CONDENAÇÃO ATUALIZADO
        'A93:V93',  # VALOR DOS HONORÁRIOS DE SUCUMBÊNCIA 5%
        'A94:V94'   # VALOR TOTAL DEVIDO PELA CONDENAÇÃO
    ]
    
    # Aplicar todas as mesclagens
    for cell_range in merged_cells_ranges:
        ws.merge_cells(cell_range)
        # Aplicar alinhamento central às células mescladas
        top_left_cell = ws[cell_range.split(':')[0]]
        top_left_cell.alignment = center_alignment
    
    # Adicionar textos finais nas células mescladas do rodapé
    ws['A92'] = 'VALOR DA CONDENAÇÃO ATUALIZADO'
    ws['A92'].font = header_font
    ws['A92'].border = thin_border
    
    ws['A93'] = 'VALOR DOS HONORÁRIOS DE SUCUMBÊNCIA 5%'
    ws['A93'].font = normal_font
    ws['A93'].border = thin_border
    
    ws['A94'] = 'VALOR TOTAL DEVIDO PELA CONDENAÇÃO'
    ws['A94'].font = header_font
    ws['A94'].border = thin_border
    
    # === APLICAR CORES ESPECÍFICAS ===
    # Cores azuis identificadas no arquivo original (linhas 1-4)
    blue_font_cells = [
        'E1', 'F1', 'G1', 'H1', 'L1',  # Linha 1
        'E2', 'F2', 'G2', 'H2', 'J2',  # Linha 2
        'E3', 'F3', 'G3', 'H3', 'J3',  # Linha 3
        'B4', 'C4', 'D4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'P4', 'Q4', 'R4', 'S4', 'T4', 'U4', 'V4', 'W4'  # Linha 4
    ]
    
    for cell_ref in blue_font_cells:
        cell = ws[cell_ref]
        cell.font = blue_font
        cell.border = thin_border
    
    # Fundo cinza para células R7 e R9
    ws['R7'].fill = gray_background
    ws['R9'].fill = gray_background
    
    # Fundo azul claro para as colunas de cálculo (R, T, V, W) nas linhas de dados
    for row in range(10, 92):  # Linhas de dados
        for col in ['R', 'T', 'V', 'W']:
            cell = ws[f'{col}{row}']
            cell.fill = light_blue_background
            cell.border = thin_border
            cell.font = normal_font
    for row in range(1, 95):  # Até linha 94 como no original
        for col in range(1, 25):  # Até coluna X (24)
            cell = ws.cell(row=row, column=col)
            if cell.border.left.style is None:  # Se ainda não tem borda
                cell.border = thin_border
                if cell.font.name is None:  # Se ainda não tem fonte
                    cell.font = normal_font
    
    # === AJUSTAR LARGURAS DAS COLUNAS ===
    column_widths = {
        'A': 10,  # Mês/Referência
        'B': 8,   # Nível  
        'C': 10,  # Vencimento Pago
        'D': 10,  # Vencimento Devido
        'E': 10,  # Vencimento Diferença
        'F': 8,   # GAM %
        'G': 10,  # GAM Pago
        'H': 10,  # GAM Devido
        'I': 10,  # GAM Diferença
        'J': 8,   # Titulação %
        'K': 10,  # Titulação Diferença
        'L': 8,   # GCET %
        'M': 10,  # GCET Diferença
        'N': 8,   # Adic. Tempo Serv. %
        'O': 10,  # Adic. Tempo Serv. Diferença
        'P': 8,   # Férias %
        'Q': 10,  # Férias Diferença
        'R': 12,  # Diferença Total (a)
        'S': 12,  # Índice de Atualização (b)
        'T': 12,  # Valor corrigido (c = a x b)
        'U': 10,  # Id.Juros % (d)
        'V': 12,  # Juros de Mora (e = c x d)
        'W': 12,  # Principal + juros (f = c + e)
        'X': 10   # Coluna extra se necessário
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    return wb

if __name__ == "__main__":
    # Criar versão exata do arquivo original
    print("Criando template baseado no arquivo original...")
    wb1 = criar_planilha_descompressao_original()
    wb1.save('Cálculos HT 2.0 - Descompressão_Template_Original.xlsx')
    print("Arquivo criado: Cálculos HT 2.0 - Descompressão_Template_Original.xlsx")